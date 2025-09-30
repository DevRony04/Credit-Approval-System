from __future__ import annotations

from decimal import Decimal
from datetime import datetime, timedelta
from celery import shared_task
from django.conf import settings
from openpyxl import load_workbook

from .models import Customer, Loan
from .services import calculate_approved_limit, amortized_monthly_payment


@shared_task
def ingest_excel_data() -> int:
    created = 0
    data_dir = settings.DATA_DIR

    customer_file = data_dir / "customer_data.xlsx"
    if customer_file.exists():
        wb = load_workbook(filename=str(customer_file))
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))[0:]]
        for row in ws.iter_rows(min_row=2):
            row_data = {headers[i]: (row[i].value if i < len(row) else None) for i in range(len(headers))}
            phone = str(row_data.get("phone_number")) if row_data.get("phone_number") is not None else None
            if not phone:
                continue
            customer, was_created = Customer.objects.get_or_create(
                phone_number=phone,
                defaults={
                    "first_name": row_data.get("first_name") or "",
                    "last_name": row_data.get("last_name") or "",
                    "age": int(row_data.get("age") or 0),
                    "monthly_income": Decimal(str(row_data.get("monthly_income") or 0)),
                    "approved_limit": calculate_approved_limit(Decimal(str(row_data.get("monthly_income") or 0))),
                    "current_debt": Decimal(str(row_data.get("current_debt") or 0)),
                },
            )
            if was_created:
                created += 1

    loan_file = data_dir / "loan_data.xlsx"
    if loan_file.exists():
        wb = load_workbook(filename=str(loan_file))
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))[0:]]
        for row in ws.iter_rows(min_row=2):
            row_data = {headers[i]: (row[i].value if i < len(row) else None) for i in range(len(headers))}
            phone = str(row_data.get("phone_number")) if row_data.get("phone_number") is not None else None
            if not phone:
                continue
            try:
                customer = Customer.objects.get(phone_number=phone)
            except Customer.DoesNotExist:
                continue
            tenure = int(row_data.get("tenure") or 0)
            rate = Decimal(str(row_data.get("interest_rate") or 0))
            amount = Decimal(str(row_data.get("loan_amount") or 0))
            start_value = row_data.get("start_date")
            if isinstance(start_value, datetime):
                start = start_value.date()
            else:
                # fallback: today
                start = datetime.utcnow().date()
            end = start + timedelta(days=30 * tenure)
            monthly = amortized_monthly_payment(amount, rate, tenure)
            Loan.objects.get_or_create(
                customer=customer,
                loan_amount=amount,
                tenure=tenure,
                interest_rate=rate,
                monthly_repayment=monthly,
                start_date=start,
                end_date=end,
                defaults={"emis_paid_on_time": int(row_data.get("emis_paid_on_time") or 0)},
            )
    return created
